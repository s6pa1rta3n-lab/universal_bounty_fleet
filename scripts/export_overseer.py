#!/usr/bin/env python3
"""Turn the cleaned overseer workbook into the JSON the Fleet Console serves."""

from __future__ import annotations

import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

SRC = Path("/Users/ankurthakur/Downloads/bounty-operations-overseer-clean.xlsx")
DST = Path(__file__).resolve().parents[1] / "app" / "data" / "overseer.json"


def _iso(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _num(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    sprint_rows = list(wb["This sprint"].iter_rows(min_row=2, values_only=True))
    prs = []
    for row in sprint_rows:
        if not row or not row[2]:
            continue
        prs.append(
            {
                "wait_on": row[0],
                "outcome": row[1],
                "repo": row[2],
                "number": str(row[3]) if row[3] is not None else None,
                "title": row[4],
                "url": row[5],
                "opened": _iso(row[6]),
                "closed": _iso(row[7]),
                "hours_open": _num(row[8]),
            }
        )

    repos = []
    for row in wb["By repo"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        repos.append(
            {
                "repo": row[0],
                "opened": int(row[1] or 0),
                "waiting": int(row[2] or 0),
                "merged": int(row[3] or 0),
                "closed": int(row[4] or 0),
                "merge_rate": row[5],
            }
        )

    claims = []
    for row in wb["Claims"].iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        claims.append(
            {
                "wait_on": row[0],
                "repo": row[1],
                "number": str(row[2]) if row[2] is not None else None,
                "title": row[3],
                "url": row[4],
                "platform": row[5],
                "status": row[6],
                "payout_usd": _num(row[7]) or 0,
                "payout_note": row[8],
            }
        )

    archive = []
    for row in wb["Prior history (parked)"].iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        archive.append(
            {
                "why_parked": row[0],
                "state": row[1],
                "repo": row[2],
                "number": str(row[3]) if row[3] is not None else None,
                "title": row[4],
                "url": row[5],
                "opened": _iso(row[6]),
            }
        )
    wb.close()

    days = Counter(pr["opened"] for pr in prs if pr.get("opened"))
    payload = {
        "meta": {
            "source": "bounty-operations-overseer-clean.xlsx",
            "snapshot": "2026-08-27",
            "window": "2026-08-24 → 2026-08-28",
            "rule": "Most open work is waiting on a human maintainer. That is not unfinished agent work.",
            "money": "Claim payouts are unknown. Treat claims as pipeline, not receivables.",
        },
        "sprint": {
            "opened": len(prs),
            "waiting": sum(1 for pr in prs if pr["outcome"] == "Waiting on human"),
            "merged": sum(1 for pr in prs if pr["outcome"] == "Merged"),
            "closed": sum(1 for pr in prs if pr["outcome"] == "Closed, not merged"),
            "days": [{"day": day, "opened": count} for day, count in sorted(days.items())],
            "repos": repos,
            "prs": prs,
        },
        "claims": claims,
        "archive": archive,
    }
    DST.parent.mkdir(parents=True, exist_ok=True)
    DST.write_text(json.dumps(payload, indent=2) + "\n")
    print(f"wrote {DST} prs={len(prs)} repos={len(repos)} claims={len(claims)} archive={len(archive)}")


if __name__ == "__main__":
    main()
